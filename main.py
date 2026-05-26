from flask import Flask, request, jsonify
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from dotenv import load_dotenv
import json

# Cargar variables de entorno (solo si existe .env local)
if os.path.exists('.env'):
    load_dotenv()

app = Flask(__name__)

# Inicializar cliente de Claude
# Inicializar cliente de Claude (API Key viene de variable de entorno)
api_key = os.getenv('ANTHROPIC_API_KEY')
if not api_key:
    raise ValueError("ANTHROPIC_API_KEY no está configurada")
client = Anthropic(api_key=api_key)

# ═══════════════════════════════════════════════════════════════
# WEBHOOK DE SPICEWORKS
# ═══════════════════════════════════════════════════════════════

@app.route('/webhook/spiceworks', methods=['POST'])
def webhook_spiceworks():
    """
    Recibe webhook de Spiceworks cuando se crea un ticket
    """
    try:
        data = request.json
        
        # Extraer datos del ticket
        ticket_id = data.get('id') or data.get('ticket_id') or 'unknown'
        descripcion = data.get('description') or data.get('body') or ''
        asunto = data.get('subject') or data.get('title') or ''
        
        print(f"\n{'='*60}")
        print(f"🔔 WEBHOOK RECIBIDO")
        print(f"{'='*60}")
        print(f"Ticket ID: {ticket_id}")
        print(f"Asunto: {asunto}")
        print(f"Descripción: {descripcion[:100]}...")
        print(f"{'='*60}\n")
        
        # Procesar con Claude
        resultado = procesar_ticket_con_claude(asunto, descripcion, ticket_id)
        
        return {
            'status': 'success',
            'message': 'Ticket procesado',
            'resultado': resultado
        }, 200
        
    except Exception as e:
        print(f"❌ Error en webhook: {str(e)}")
        return {'status': 'error', 'error': str(e)}, 500


# ═══════════════════════════════════════════════════════════════
# PROCESAR CON CLAUDE
# ═══════════════════════════════════════════════════════════════

def procesar_ticket_con_claude(asunto, descripcion, ticket_id):
    """
    Claude analiza el ticket y ejecuta acciones
    """
    
    # Construir el prompt para Claude
    prompt = f"""
Eres un asistente que procesa solicitudes de crear usuarios.

Analiza esta solicitud de ticket:
Asunto: {asunto}
Descripción: {descripcion}

Si detectas una solicitud para CREAR USUARIOS, extrae la información en este formato JSON:
{{
  "accion": "crear_usuarios",
  "usuarios": [
    {{"nombre": "Nombre Completo", "email": "email@empresa.pe"}},
    {{"nombre": "Otro Nombre", "email": "otro@empresa.pe"}}
  ]
}}

Si NO es una solicitud de crear usuarios, responde:
{{"accion": "otro", "mensaje": "No es una solicitud de crear usuarios"}}

IMPORTANTE: Responde SOLO con JSON válido, sin explicaciones adicionales.
"""
    
    print(f"🤖 Claude analizando ticket...")
    
    # Llamar a Claude
    response = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=1000,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )
    
    respuesta_texto = response.content[0].text
    print(f"Claude respondió: {respuesta_texto}")
    
    # Parsear respuesta JSON
    try:
        data_json = json.loads(respuesta_texto)
    except json.JSONDecodeError:
        print("❌ Claude no respondió con JSON válido")
        return {"status": "error", "error": "Invalid JSON from Claude"}
    
    # Ejecutar acción
    if data_json.get("accion") == "crear_usuarios":
        usuarios = data_json.get("usuarios", [])
        print(f"\n✓ Claude detectó: Crear {len(usuarios)} usuarios")
        
        # Crear Excel
        resultado = crear_excel_usuarios(usuarios, ticket_id)
        return resultado
    
    return {"status": "ok", "mensaje": "Ticket procesado pero no es creación de usuarios"}


# ═══════════════════════════════════════════════════════════════
# CREAR EXCEL CON DATOS DE USUARIOS
# ═══════════════════════════════════════════════════════════════

def crear_excel_usuarios(usuarios, ticket_id):
    """
    Crea archivo Excel con datos de los usuarios a crear
    """
    
    print(f"\n📊 Creando archivo Excel...")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Nombre", "Email", "Fecha Creación", "Estado"]
    ws.append(headers)
    
    # Aplicar estilos a headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Agregar datos de usuarios
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    for usuario in usuarios:
        nombre = usuario.get("nombre", "N/A")
        email = usuario.get("email", "N/A")
        
        ws.append([
            nombre,
            email,
            fecha_actual,
            "Pendiente creación"  # En simulación, aún no creamos en M365
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Guardar archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"usuarios_creados_{ticket_id}_{timestamp}.xlsx"
    filepath = f"/tmp/{filename}"  # En Mac, /tmp es temporal
    
    wb.save(filepath)
    
    print(f"✓ Archivo creado: {filename}")
    print(f"  Ubicación: {filepath}")
    
    return {
        "status": "success",
        "archivo": filename,
        "ruta": filepath,
        "usuarios_procesados": len(usuarios),
        "detalles": [f"{u['nombre']} ({u['email']})" for u in usuarios]
    }


# ═══════════════════════════════════════════════════════════════
# RUTA DE PRUEBA
# ═══════════════════════════════════════════════════════════════

@app.route('/', methods=['GET'])
def home():
    """Ruta principal para verificar que servidor está online"""
    return {
        'status': 'online',
        'mensaje': '✓ Servidor Driftnode running',
        'webhook': 'POST /webhook/spiceworks'
    }, 200


@app.route('/test', methods=['POST'])
def test():
    """Ruta de prueba sin necesidad de Spiceworks"""
    data = request.json
    asunto = data.get('asunto', 'Test: Crear usuarios')
    descripcion = data.get('descripcion', 'Crear usuarios: Juan García (juan@test.pe), María López (maria@test.pe)')
    
    print(f"\n🧪 TEST RECIBIDO")
    resultado = procesar_ticket_con_claude(asunto, descripcion, "test-001")
    
    return resultado, 200


# ═══════════════════════════════════════════════════════════════
# INICIAR SERVIDOR
# ═══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 SERVIDOR DRIFTNODE INICIANDO")
    print("="*60)
    print("URL: http://localhost:5000")
    print("Webhook: POST /webhook/spiceworks")
    print("Test: POST /test")
    print("="*60 + "\n")
    
    app.run(debug=True, host='localhost', port=5000)
